from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import LOGO_PATH


def _fmt(value: float | None) -> str | float:
    return "" if value is None else round(value, 2)


def _detail_row_values(row: dict) -> list:
    values = [row["kelas"], row["no"], row["nama"]]
    for score in row["materi"]:
        values.extend([_fmt(score["hafalan"]), _fmt(score["praktik"])])
    values.append(_fmt(row["rata_akhir"]))
    return values


def export_laporan_excel(path: str | Path, metadata: dict[str, Any], rekap_list, detail_grid: dict | None = None) -> str:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap"

    if LOGO_PATH.exists():
        try:
            logo = XLImage(str(LOGO_PATH))
            logo.width = 58
            logo.height = 58
            ws.add_image(logo, "A1")
        except Exception:
            pass

    ws["B1"] = metadata.get("nama_sekolah", "")
    ws["B1"].font = Font(bold=True, size=15)
    ws["B2"] = "Laporan Penilaian Praktik Ibadah (PIB)"
    ws["B3"] = f"{metadata.get('tahun_ajaran', '-')} | {metadata.get('semester', '-')} | {metadata.get('kelas', '-')}"
    ws["B4"] = f"Tanggal: {metadata.get('tanggal_export', '-')}"

    headers = detail_grid["headers"] if detail_grid else ["Peringkat", "NIS", "Nama", "Kelas", "Rata Hafalan", "Rata Praktik", "Rata Total", "Status"]
    start_row = 6
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF4FF")

    source_rows = detail_grid["rows"] if detail_grid else rekap_list
    for row_idx, item in enumerate(source_rows, start=start_row + 1):
        if detail_grid:
            values = _detail_row_values(item)
        else:
            values = [
                item.peringkat or "",
                item.nis,
                item.nama,
                item.kelas,
                _fmt(item.rata_hafalan),
                _fmt(item.rata_praktik),
                _fmt(item.rata_total),
                item.status,
            ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)

    if detail_grid:
        ws_rekap = wb.create_sheet("Ranking")
        ranking_headers = ["Peringkat", "NIS", "Nama", "Kelas", "Rata Hafalan", "Rata Praktik", "Rata Total", "Status"]
        for col, header in enumerate(ranking_headers, start=1):
            cell = ws_rekap.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EEF4FF")
        for row_idx, item in enumerate(rekap_list, start=2):
            values = [
                item.peringkat or "",
                item.nis,
                item.nama,
                item.kelas,
                _fmt(item.rata_hafalan),
                _fmt(item.rata_praktik),
                _fmt(item.rata_total),
                item.status,
            ]
            for col_idx, value in enumerate(values, start=1):
                ws_rekap.cell(row_idx, col_idx, value)
        for col_idx in range(1, len(ranking_headers) + 1):
            letter = get_column_letter(col_idx)
            max_len = max(len(str(ws_rekap.cell(row, col_idx).value or "")) for row in range(1, ws_rekap.max_row + 1))
            ws_rekap.column_dimensions[letter].width = min(max(max_len + 3, 12), 34)

    ws.freeze_panes = "A7"
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_len = max(len(str(ws.cell(row, col_idx).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 34)

    wb.save(path)
    return str(path)
