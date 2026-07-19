from __future__ import annotations

from tkinter import filedialog, messagebox

import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import EXPORT_DIR

from ui import theme
from ui.components.widgets import (
    EmptyState,
    SectionCard,
    TableHeader,
    TableRow,
    badge,
    danger_button,
    primary_button,
    secondary_button,
    show_loading,
    show_toast,
    styled_entry,
    styled_option,
)


class SiswaView(ctk.CTkFrame):
    def __init__(self, master, repo, refresh_callback):
        super().__init__(master, fg_color=theme.BACKGROUND)
        self.repo = repo
        self.refresh_callback = refresh_callback
        self.semester = repo.get_semester_aktif()
        self.kelas_list = repo.list_kelas()
        self.kelas_map = {k["nama"]: k["id"] for k in self.kelas_list}
        self.kelas_var = ctk.StringVar(value="Semua")
        self.search_var = ctk.StringVar()
        self.editing_siswa_id: int | None = None
        self.grid_columnconfigure(0, weight=1)

        hero = ctk.CTkFrame(self, fg_color=theme.SURFACE_ACCENT, border_color=theme.BORDER, border_width=1, corner_radius=12)
        hero.grid(row=0, column=0, sticky="ew", pady=(0, 18))
        hero.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(hero, text="Manajemen Data Siswa", font=(theme.FONT, 26, "bold"), text_color=theme.TEXT).grid(
            row=0, column=0, sticky="w", padx=20, pady=(18, 4)
        )
        ctk.CTkLabel(
            hero,
            text="Tambah siswa, filter per kelas, dan pantau status penilaiannya.",
            font=(theme.FONT, 14),
            text_color=theme.TEXT_MUTED,
        ).grid(row=1, column=0, sticky="w", padx=20, pady=(0, 18))
        badge(hero, f"{len(self.repo.list_siswa(self.semester['id'])) if self.semester else 0} siswa", "info").grid(
            row=0, column=1, rowspan=2, padx=20, pady=20
        )

        if not self.semester:
            EmptyState(self, "Belum ada semester aktif", "Buat periode aktif sebelum menambah siswa.").grid(row=1, column=0, sticky="ew")
            return

        self._form().grid(row=1, column=0, sticky="ew", pady=(0, 18))
        self._import_card().grid(row=2, column=0, sticky="ew", pady=(0, 18))
        self.table_card = self._table()
        self.table_card.grid(row=3, column=0, sticky="ew")

    def _form(self):
        card = SectionCard(self, "Tambah Siswa", "Masukkan data siswa untuk semester aktif.")
        card.body.grid_columnconfigure(1, weight=1)
        self.nis_var = ctk.StringVar()
        self.nama_var = ctk.StringVar()
        self.kelas_input_var = ctk.StringVar(value=self.kelas_list[0]["nama"] if self.kelas_list else "")
        form = ctk.CTkFrame(card.body, fg_color=theme.SURFACE)
        form.grid(row=0, column=0, columnspan=4, sticky="ew", padx=16, pady=(4, 16))
        form.grid_columnconfigure(1, weight=1)

        self._field_label(form, "NIS").grid(row=0, column=0, sticky="w", pady=(0, 4))
        self._field_label(form, "Nama Siswa").grid(row=0, column=1, sticky="w", padx=(12, 0), pady=(0, 4))
        self._field_label(form, "Kelas").grid(row=0, column=2, sticky="w", padx=(12, 0), pady=(0, 4))

        styled_entry(form, self.nis_var, "001", width=150).grid(row=1, column=0, sticky="ew")
        styled_entry(form, self.nama_var, "Nama lengkap siswa").grid(row=1, column=1, sticky="ew", padx=(12, 0))
        styled_option(form, list(self.kelas_map) or ["-"], self.kelas_input_var, width=130).grid(row=1, column=2, sticky="ew", padx=(12, 0))
        self.save_button = primary_button(form, "Tambah Siswa", command=self._save_siswa, width=130)
        self.save_button.grid(
            row=1, column=3, padx=(12, 0)
        )
        self.cancel_edit_button = secondary_button(form, "Batal", command=self._cancel_edit_siswa, width=80)
        self.cancel_edit_button.grid(row=1, column=4, padx=(8, 0))
        self.cancel_edit_button.grid_remove()
        return card

    def _field_label(self, parent, text: str):
        return ctk.CTkLabel(
            parent,
            text=text,
            font=(theme.FONT, 13, "bold"),
            text_color=theme.TEXT,
            anchor="w",
        )

    def _import_card(self):
        card = SectionCard(
            self,
            "Import Excel",
            "Unduh template, isi NIS, Nama, dan Kelas, lalu impor kembali ke aplikasi.",
        )
        row = ctk.CTkFrame(card.body, fg_color=theme.SURFACE)
        row.pack(fill="x", padx=16, pady=(2, 14))
        ctk.CTkLabel(
            row,
            text="Format wajib: NIS | Nama | Kelas. Nama kelas harus sudah terdaftar di menu Periode.",
            font=(theme.FONT, 13),
            text_color=theme.TEXT_MUTED,
            anchor="w",
            wraplength=720,
            justify="left",
        ).pack(side="left", fill="x", expand=True)
        secondary_button(row, "Unduh Template", command=self._download_template, width=150).pack(side="right", padx=(10, 0))
        primary_button(row, "Import Excel", command=self._import_excel, width=130).pack(side="right", padx=(10, 0))
        return card

    def _table(self):
        card = SectionCard(self, "Daftar Siswa", "Gunakan pencarian dan filter untuk menemukan siswa dengan cepat.")
        filter_row = ctk.CTkFrame(card.body, fg_color=theme.SURFACE)
        filter_row.pack(fill="x", padx=16, pady=14)
        ctk.CTkLabel(filter_row, text="Filter Kelas:", font=(theme.FONT, 13, "bold"), text_color=theme.TEXT_MUTED).pack(side="left")
        values = ["Semua", *list(self.kelas_map)]
        styled_option(filter_row, values, self.kelas_var, command=lambda _: self._render_rows(), width=150).pack(side="left", padx=12)
        search = styled_entry(filter_row, self.search_var, "Cari NIS atau nama...", width=280)
        search.pack(side="left", padx=(8, 0))
        self.search_var.trace_add("write", lambda *_: self._render_rows())

        header = TableHeader(card.body, [("No", 58), ("NIS", 120), ("Nama Siswa", None), ("Kelas", 90), ("Status", 110), ("Aksi", 250)])
        header.pack(fill="x")
        self.rows_frame = ctk.CTkFrame(card.body, fg_color=theme.SURFACE)
        self.rows_frame.pack(fill="x")
        self._render_rows()
        return card

    def _render_rows(self):
        if not hasattr(self, "rows_frame"):
            return
        for child in self.rows_frame.winfo_children():
            child.destroy()

        kelas_id = self.kelas_map.get(self.kelas_var.get())
        siswa = self.repo.list_siswa(self.semester["id"], kelas_id)
        query = self.search_var.get().strip().lower()
        if query:
            siswa = [s for s in siswa if query in s["nis"].lower() or query in s["nama"].lower()]
        rekap_by_id = {r.siswa_id: r for r in self.repo.get_rekap_semester(self.semester["id"])}
        for idx, item in enumerate(siswa, start=1):
            row = TableRow(self.rows_frame, idx, height=56)
            row.pack(fill="x", pady=2)
            for value, width, expand in [
                (str(idx), 58, False),
                (item["nis"], 120, False),
                (item["nama"], 1, True),
                (item["kelas"], 90, False),
            ]:
                kwargs = {} if expand else {"width": width}
                ctk.CTkLabel(row, text=value, anchor="w", font=(theme.FONT, 13), text_color=theme.TEXT, **kwargs).pack(
                    side="left", fill="x", expand=expand, padx=18, pady=10
                )
            status = rekap_by_id.get(item["id"]).status if item["id"] in rekap_by_id else "belum"
            badge(row, status, status).pack(side="left", padx=18)
            secondary_button(row, "Edit", command=lambda data=item: self._start_edit_siswa(data), width=72).pack(side="left", padx=(0, 8), pady=8)
            secondary_button(row, "Nonaktif", command=lambda siswa_id=item["id"]: self._deactivate(siswa_id), width=88).pack(side="left", padx=(0, 8), pady=8)
            danger_button(row, "Hapus", command=lambda siswa_id=item["id"]: self._delete_siswa(siswa_id), width=76).pack(side="left", padx=(0, 12), pady=8)
        if not siswa:
            EmptyState(self.rows_frame, "Data siswa tidak ditemukan", "Coba ubah kata kunci pencarian atau filter kelas.").pack(
                fill="x", padx=14, pady=14
            )

    def _save_siswa(self):
        kelas_id = self.kelas_map.get(self.kelas_input_var.get())
        if not kelas_id:
            messagebox.showerror("Kelas belum ada", "Tambahkan kelas terlebih dahulu.")
            return
        try:
            data = {
                "nis": self.nis_var.get().strip(),
                "nama": self.nama_var.get().strip(),
                "kelas_id": kelas_id,
                "semester_id": self.semester["id"],
            }
            if self.editing_siswa_id is None:
                self.repo.create_siswa(data)
                show_toast(self, "Siswa ditambahkan", f"{data['nama']} masuk ke kelas {self.kelas_input_var.get()}.")
            else:
                self.repo.update_siswa(self.editing_siswa_id, data)
                show_toast(self, "Siswa diperbarui", f"Data {data['nama']} berhasil diedit.", "info")
            self._cancel_edit_siswa(reset_only=True)
            self.refresh_callback()
        except Exception as exc:
            messagebox.showerror("Gagal menyimpan siswa", str(exc))

    def _start_edit_siswa(self, data: dict):
        self.editing_siswa_id = data["id"]
        self.nis_var.set(data["nis"])
        self.nama_var.set(data["nama"])
        self.kelas_input_var.set(data["kelas"])
        self.save_button.configure(text="Simpan Edit")
        self.cancel_edit_button.grid()
        show_toast(self, "Mode edit siswa", f"Ubah data {data['nama']} lalu simpan.", "info")

    def _cancel_edit_siswa(self, reset_only: bool = False):
        self.editing_siswa_id = None
        self.nis_var.set("")
        self.nama_var.set("")
        if self.kelas_list:
            self.kelas_input_var.set(self.kelas_list[0]["nama"])
        self.save_button.configure(text="Tambah Siswa")
        self.cancel_edit_button.grid_remove()
        if not reset_only:
            show_toast(self, "Edit dibatalkan", "Form kembali ke mode tambah.", "info")

    def _download_template(self):
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        default_path = EXPORT_DIR / "template_import_siswa.xlsx"
        path = filedialog.asksaveasfilename(
            title="Simpan Template Import Siswa",
            defaultextension=".xlsx",
            initialdir=EXPORT_DIR,
            initialfile=default_path.name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Import Siswa"
        headers = ["NIS", "Nama", "Kelas"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="111C2D")
            cell.fill = PatternFill("solid", fgColor="EEF4FF")

        contoh_kelas = self.kelas_list[0]["nama"] if self.kelas_list else "7A"
        examples = [
            ["001", "Ahmad Fauzi", contoh_kelas],
            ["002", "Siti Aminah", contoh_kelas],
        ]
        for row_idx, row in enumerate(examples, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        guide = wb.create_sheet("Panduan")
        guide["A1"] = "Panduan Import Siswa"
        guide["A1"].font = Font(bold=True, size=14)
        guide["A3"] = "1. Isi data pada sheet 'Import Siswa'."
        guide["A4"] = "2. Kolom wajib: NIS, Nama, Kelas."
        guide["A5"] = "3. Nama kelas harus sama dengan kelas yang sudah dibuat di aplikasi."
        guide["A6"] = "4. Jangan mengubah nama header kolom."
        guide["A8"] = "Kelas tersedia:"
        for idx, kelas in enumerate(self.kelas_list, start=9):
            guide.cell(row=idx, column=1, value=kelas["nama"])

        for sheet in [ws, guide]:
            for col_idx in range(1, 5):
                sheet.column_dimensions[get_column_letter(col_idx)].width = 24
        loading = show_loading(self, "Membuat template", "Menyiapkan file Excel import siswa...")
        try:
            wb.save(path)
        finally:
            loading.close()
        show_toast(self, "Template dibuat", f"Tersimpan di {path}", "success")

    def _import_excel(self):
        path = filedialog.askopenfilename(
            title="Pilih File Excel Siswa",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not path:
            return

        try:
            loading = show_loading(self, "Mengimpor siswa", "Membaca file Excel dan memvalidasi kelas...")
            wb = load_workbook(path)
            ws = wb["Import Siswa"] if "Import Siswa" in wb.sheetnames else wb.active
            header_map = {
                str(cell.value).strip().lower(): idx
                for idx, cell in enumerate(ws[1], start=1)
                if cell.value is not None
            }
            required = {"nis", "nama", "kelas"}
            if not required.issubset(header_map):
                raise ValueError("Header wajib harus ada: NIS, Nama, Kelas")

            imported = 0
            skipped: list[str] = []
            for row_idx in range(2, ws.max_row + 1):
                nis = str(ws.cell(row_idx, header_map["nis"]).value or "").strip()
                nama = str(ws.cell(row_idx, header_map["nama"]).value or "").strip()
                kelas_name = str(ws.cell(row_idx, header_map["kelas"]).value or "").strip().upper()
                if not nis and not nama and not kelas_name:
                    continue
                if not nis or not nama or not kelas_name:
                    skipped.append(f"Baris {row_idx}: data belum lengkap")
                    continue
                kelas_id = self.kelas_map.get(kelas_name)
                if not kelas_id:
                    skipped.append(f"Baris {row_idx}: kelas '{kelas_name}' belum terdaftar")
                    continue
                try:
                    self.repo.create_siswa(
                        {
                            "nis": nis,
                            "nama": nama,
                            "kelas_id": kelas_id,
                            "semester_id": self.semester["id"],
                        }
                    )
                    imported += 1
                except Exception as exc:
                    skipped.append(f"Baris {row_idx}: {exc}")

            detail = "" if not skipped else "\n\nTidak diimpor:\n" + "\n".join(skipped[:10])
            if len(skipped) > 10:
                detail += f"\n...dan {len(skipped) - 10} baris lainnya."
            loading.close()
            if skipped:
                show_toast(self, "Import selesai sebagian", f"{imported} siswa masuk, {len(skipped)} baris dilewati.", "warning")
                messagebox.showinfo("Detail import", f"{imported} siswa berhasil diimpor.{detail}")
            else:
                show_toast(self, "Import selesai", f"{imported} siswa berhasil ditambahkan.")
            self.refresh_callback()
        except Exception as exc:
            if "loading" in locals() and loading.winfo_exists():
                loading.close()
            messagebox.showerror("Import gagal", str(exc))

    def _deactivate(self, siswa_id: int):
        if not messagebox.askyesno("Nonaktifkan siswa", "Siswa akan disembunyikan dari daftar aktif. Lanjutkan?"):
            return
        self.repo.deactivate_siswa(siswa_id)
        show_toast(self, "Siswa dinonaktifkan", "Data siswa disembunyikan dari daftar aktif.", "warning")
        self._render_rows()

    def _delete_siswa(self, siswa_id: int):
        if not messagebox.askyesno("Hapus siswa", "Siswa dan seluruh nilai terkait akan dihapus permanen. Lanjutkan?"):
            return
        self.repo.delete_siswa(siswa_id)
        show_toast(self, "Siswa dihapus", "Data siswa dan nilai terkait sudah dihapus.", "warning")
        self._render_rows()
